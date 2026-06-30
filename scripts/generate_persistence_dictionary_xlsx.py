from __future__ import annotations

from pathlib import Path
import sys


ROWS = [
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "data_sources",
        "module": "datasource, catalog, scanner",
        "purpose": "Cadastro de conexões/fontes de dados, credenciais cifradas e escopo de scan.",
        "orm_model": "DataSource",
        "orm_file": "backend/app/models/catalog.py",
        "migration": "55da547ae3ae_baseline.py; e6d4a1b9c2f3_add_connection_config_to_datasources.py",
        "writers": "backend/app/api/datasource.py; backend/app/services/datasource.py",
        "readers": "catalog, dashboard, search, scan, privacy-access, certification",
        "trigger_endpoint": "POST/PUT/DELETE /api/v1/datasources/*",
        "relationships": "1:N com databases; 1:N com scan_runs",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "databases",
        "module": "catalog, scanner",
        "purpose": "Bancos descobertos/catalogados por datasource.",
        "orm_model": "Database",
        "orm_file": "backend/app/models/catalog.py",
        "migration": "55da547ae3ae_baseline.py",
        "writers": "backend/app/services/scanner.py",
        "readers": "catalog, dashboard, certification, privacy-access",
        "trigger_endpoint": "POST /api/v1/scan-runs/datasource/{id}",
        "relationships": "N:1 com data_sources; 1:N com schemas",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "schemas",
        "module": "catalog, scanner",
        "purpose": "Schemas descobertos/catalogados.",
        "orm_model": "Schema",
        "orm_file": "backend/app/models/catalog.py",
        "migration": "55da547ae3ae_baseline.py",
        "writers": "backend/app/services/scanner.py",
        "readers": "catalog, dashboard, certification, privacy-access, search",
        "trigger_endpoint": "POST /api/v1/scan-runs/datasource/{id}",
        "relationships": "N:1 com databases; 1:N com tables",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "tables",
        "module": "catalog, explorer, ownership, certification, privacy/access, dashboard",
        "purpose": "Ativo central do catálogo com metadados manuais e de origem, ownership, certificação e privacidade.",
        "orm_model": "TableEntity",
        "orm_file": "backend/app/models/catalog.py",
        "migration": "55da547ae3ae_baseline.py; 3c1f0b7d2e4a_add_data_owners.py; a31c5d9f4b22_add_table_certification_fields.py; c4e7a2b1d9f0_add_certification_badges_column.py; f7c1d2e3a4b5_add_table_privacy_access_fields.py",
        "writers": "backend/app/services/scanner.py; backend/app/api/catalog.py; backend/app/api/table_metadata.py; backend/app/api/certification.py; backend/app/api/privacy_access.py",
        "readers": "praticamente todo o sistema",
        "trigger_endpoint": "POST /api/v1/scan-runs/datasource/{id}; PATCH /api/v1/catalog/tables/{id}; PATCH /api/v1/tables/{id}; PATCH /api/v1/certification/tables/{id}; PATCH /api/v1/privacy-access/tables/{id}",
        "relationships": "N:1 com schemas; N:1 com data_owners; 1:N com columns; referenciado por dq, lineage, tags e glossary",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "columns",
        "module": "catalog, dictionary, explorer, dq",
        "purpose": "Metadados de colunas e dicionário de dados.",
        "orm_model": "ColumnEntity",
        "orm_file": "backend/app/models/catalog.py",
        "migration": "55da547ae3ae_baseline.py; 9b7e3c2d1f44_add_column_dictionary_fields.py",
        "writers": "backend/app/services/scanner.py; backend/app/services/column_dictionary_spreadsheet.py",
        "readers": "catalog, explorer, certification, dashboard, dq",
        "trigger_endpoint": "POST /api/v1/scan-runs/datasource/{id}; POST /api/v1/catalog/column-dictionary/import",
        "relationships": "N:1 com tables",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "data_owners",
        "module": "ownership",
        "purpose": "Cadastro funcional de Data Owners.",
        "orm_model": "DataOwner",
        "orm_file": "backend/app/models/catalog.py",
        "migration": "3c1f0b7d2e4a_add_data_owners.py",
        "writers": "backend/app/api/data_owners.py",
        "readers": "tables, dashboard, privacy-access, certification",
        "trigger_endpoint": "CRUD /api/v1/data-owners/*",
        "relationships": "1:N com tables",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "scan_runs",
        "module": "scanner",
        "purpose": "Histórico de execuções de scan por datasource.",
        "orm_model": "ScanRun",
        "orm_file": "backend/app/models/scan.py",
        "migration": "55da547ae3ae_baseline.py",
        "writers": "backend/app/services/scanner.py",
        "readers": "scan, debug operacional",
        "trigger_endpoint": "POST /api/v1/scan-runs/datasource/{id}",
        "relationships": "N:1 com data_sources; 1:N com scan_snapshots; 1:N com scan_diffs",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "scan_snapshots",
        "module": "scanner",
        "purpose": "Snapshot serializado do resultado de um scan.",
        "orm_model": "ScanSnapshot",
        "orm_file": "backend/app/models/scan.py",
        "migration": "55da547ae3ae_baseline.py",
        "writers": "backend/app/services/scanner.py",
        "readers": "histórico/compare de scan",
        "trigger_endpoint": "POST /api/v1/scan-runs/datasource/{id}",
        "relationships": "N:1 com scan_runs",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "scan_diffs",
        "module": "scanner",
        "purpose": "Diferenças detectadas entre scans.",
        "orm_model": "ScanDiff",
        "orm_file": "backend/app/models/scan.py",
        "migration": "55da547ae3ae_baseline.py",
        "writers": "backend/app/services/scanner.py",
        "readers": "scan, auditoria de mudanças",
        "trigger_endpoint": "POST /api/v1/scan-runs/datasource/{id}",
        "relationships": "N:1 com scan_runs",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "tags",
        "module": "tags",
        "purpose": "Catálogo de tags.",
        "orm_model": "Tag",
        "orm_file": "backend/app/models/tag.py",
        "migration": "55da547ae3ae_baseline.py; 6f4d2b8c1a90_expand_tags_for_taxonomy_and_spreadsheets.py",
        "writers": "backend/app/api/tags.py; backend/app/services/tag_spreadsheet.py; backend/app/api/import_export.py",
        "readers": "tags, dashboard, certification, search",
        "trigger_endpoint": "CRUD /api/v1/tags/*; import/export de tags; POST /api/v1/io/import",
        "relationships": "1:N com tag_assignments",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "tag_assignments",
        "module": "tags",
        "purpose": "Vínculo de tags com entidades, principalmente table.",
        "orm_model": "TagAssignment",
        "orm_file": "backend/app/models/tag.py",
        "migration": "55da547ae3ae_baseline.py",
        "writers": "backend/app/api/tags.py; backend/app/api/table_metadata.py",
        "readers": "dashboard, certification, search, explorer",
        "trigger_endpoint": "PUT /api/v1/tables/{id}/tags; endpoints de assign/unassign de tags",
        "relationships": "N:1 com tags; referência lógica por entity_type/entity_id",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "glossary_terms",
        "module": "glossary",
        "purpose": "Glossário corporativo.",
        "orm_model": "GlossaryTerm",
        "orm_file": "backend/app/models/glossary.py",
        "migration": "55da547ae3ae_baseline.py; 8a2d4f1c6b77_expand_glossary_for_taxonomy_and_spreadsheets.py",
        "writers": "backend/app/api/glossary.py; backend/app/services/glossary_spreadsheet.py; backend/app/api/import_export.py",
        "readers": "glossary, dashboard, certification, search",
        "trigger_endpoint": "CRUD /api/v1/glossary/*; import/export de glossary; POST /api/v1/io/import",
        "relationships": "1:N com glossary_assignments",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "glossary_assignments",
        "module": "glossary",
        "purpose": "Vínculo entre termos e entidades, principalmente table.",
        "orm_model": "GlossaryAssignment",
        "orm_file": "backend/app/models/glossary.py",
        "migration": "55da547ae3ae_baseline.py",
        "writers": "backend/app/api/glossary.py; backend/app/api/table_metadata.py",
        "readers": "dashboard, certification, search, explorer",
        "trigger_endpoint": "PUT /api/v1/tables/{id}/glossary-terms; endpoints de assign/unassign de glossary",
        "relationships": "N:1 com glossary_terms; referência lógica por entity_type/entity_id",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "dq_runs",
        "module": "data quality",
        "purpose": "Execução de profiling/rules, status, engine, logs e payload estruturado.",
        "orm_model": "DQRun",
        "orm_file": "backend/app/models/dq.py",
        "migration": "e0a4d2c3b901_add_dq_observability_tables.py; 1d6c80ac7e58_add_spark_fields_to_dq_runs.py; 85e30b8d49cd_add_execution_engine_to_dq_runs.py; 7c5bf86829b1_add_schema_scope_to_dq_runs.py; d4b9e6a1c2f0_store_dq_profiling_payload_in_postgres.py",
        "writers": "backend/app/services/dq.py; backend/app/services/dq_spark.py",
        "readers": "dq, dashboard, certification, home",
        "trigger_endpoint": "POST/launch endpoints em /api/v1/dq/*",
        "relationships": "1:N com dq_table_metrics; autorreferência pai/filho",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "dq_table_metrics",
        "module": "data quality",
        "purpose": "Métricas agregadas por tabela em cada run.",
        "orm_model": "DQTableMetric",
        "orm_file": "backend/app/models/dq.py",
        "migration": "e0a4d2c3b901_add_dq_observability_tables.py; d4b9e6a1c2f0_store_dq_profiling_payload_in_postgres.py",
        "writers": "backend/app/services/dq.py; backend/app/services/dq_spark.py",
        "readers": "dq, dashboard, certification, home",
        "trigger_endpoint": "POST/launch endpoints em /api/v1/dq/*",
        "relationships": "N:1 com dq_runs; N:1 com tables; 1:N com dq_column_metrics",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "dq_column_metrics",
        "module": "data quality",
        "purpose": "Métricas por coluna em cada run.",
        "orm_model": "DQColumnMetric",
        "orm_file": "backend/app/models/dq.py",
        "migration": "e0a4d2c3b901_add_dq_observability_tables.py",
        "writers": "backend/app/services/dq.py; backend/app/services/dq_spark.py",
        "readers": "dq, explorer",
        "trigger_endpoint": "POST/launch endpoints em /api/v1/dq/*",
        "relationships": "N:1 com dq_table_metrics; N:1 com columns",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "dq_rules",
        "module": "data quality",
        "purpose": "Regras de qualidade cadastradas por tabela.",
        "orm_model": "DQRule",
        "orm_file": "backend/app/models/dq.py",
        "migration": "9f2a6d1d77aa_add_dq_rules_tables.py; b742be9a32c1_ensure_dq_rules_tables_exist.py",
        "writers": "backend/app/api/dq.py",
        "readers": "dq",
        "trigger_endpoint": "CRUD /api/v1/dq/rules*",
        "relationships": "N:1 opcional com tables; 1:N com dq_rule_runs",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "dq_rule_runs",
        "module": "data quality",
        "purpose": "Resultado de execução de regra DQ.",
        "orm_model": "DQRuleRun",
        "orm_file": "backend/app/models/dq.py",
        "migration": "9f2a6d1d77aa_add_dq_rules_tables.py",
        "writers": "backend/app/services/dq.py",
        "readers": "dq, incidentes de DQ",
        "trigger_endpoint": "Execução de regras via /api/v1/dq/*",
        "relationships": "N:1 com dq_rules; opcional N:1 com dq_runs",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "dq_job_runs",
        "module": "data quality",
        "purpose": "Execução operacional de jobs Spark de DQ.",
        "orm_model": "DQJobRun",
        "orm_file": "backend/app/models/dq.py",
        "migration": "b1d3a9c2f401_add_dq_job_runs_for_spark_execution.py; 6c0464c9c46a_add_dq_run_id_to_dq_job_runs.py",
        "writers": "backend/app/services/dq_spark.py",
        "readers": "monitoramento de jobs DQ",
        "trigger_endpoint": "Execução assíncrona Spark via /api/v1/dq/*",
        "relationships": "opcional N:1 com dq_runs",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "lineage_assets",
        "module": "lineage",
        "purpose": "Ativos de linhagem normalizados: table, view, dashboard e source.",
        "orm_model": "LineageAsset",
        "orm_file": "backend/app/models/lineage.py",
        "migration": "ab39c8d4e2f1_add_manual_lineage_assets_and_relations.py",
        "writers": "backend/app/features/lineage/openlineage_sync.py; backend/app/services/lineage.py; backend/app/services/lineage_spreadsheet.py",
        "readers": "lineage, explorer",
        "trigger_endpoint": "CRUD /api/v1/lineage/assets*; import/export lineage; sync lineage",
        "relationships": "opcional N:1 com tables; self via lineage_relations",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "lineage_relations",
        "module": "lineage",
        "purpose": "Arestas de linhagem manuais, automáticas ou merged.",
        "orm_model": "LineageRelation",
        "orm_file": "backend/app/models/lineage.py",
        "migration": "ab39c8d4e2f1_add_manual_lineage_assets_and_relations.py",
        "writers": "backend/app/features/lineage/openlineage_sync.py; backend/app/services/lineage.py; backend/app/services/lineage_spreadsheet.py",
        "readers": "lineage, explorer, dashboard indireto",
        "trigger_endpoint": "CRUD /api/v1/lineage/edges*; import/export lineage; sync lineage",
        "relationships": "N:1 com lineage_assets como source/target; opcional N:1 com lineage_jobs",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "lineage_source_configs",
        "module": "lineage",
        "purpose": "Configuração de fontes automáticas de lineage internas baseadas em OpenLineage.",
        "orm_model": "LineageSourceConfig",
        "orm_file": "backend/app/models/lineage.py",
        "migration": "0b1c2d3e4f5a_add_internal_openlineage_tables.py",
        "writers": "backend/app/features/lineage/openlineage_sync.py; backend/app/api/lineage.py",
        "readers": "lineage",
        "trigger_endpoint": "POST/PATCH /api/v1/lineage/sources*",
        "relationships": "1:N com lineage_jobs; 1:N com lineage_runs",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "lineage_jobs",
        "module": "lineage",
        "purpose": "Jobs sincronizados de fonte automática de lineage.",
        "orm_model": "LineageJob",
        "orm_file": "backend/app/models/lineage.py",
        "migration": "0b1c2d3e4f5a_add_internal_openlineage_tables.py",
        "writers": "backend/app/features/lineage/openlineage_sync.py",
        "readers": "lineage, explorer",
        "trigger_endpoint": "POST /api/v1/lineage/sources/{id}/sync",
        "relationships": "N:1 com lineage_source_configs; 1:N com lineage_runs; pode ser referenciado por lineage_relations",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "lineage_runs",
        "module": "lineage",
        "purpose": "Runs de jobs sincronizados de OpenLineage.",
        "orm_model": "LineageRun",
        "orm_file": "backend/app/models/lineage.py",
        "migration": "0b1c2d3e4f5a_add_internal_openlineage_tables.py",
        "writers": "backend/app/features/lineage/openlineage_sync.py",
        "readers": "lineage, explorer",
        "trigger_endpoint": "POST /api/v1/lineage/sources/{id}/sync; POST /api/v1/lineage/tables/{id}/sync",
        "relationships": "N:1 com lineage_source_configs; N:1 com lineage_jobs",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_ops",
        "table": "incidents",
        "module": "incidents, data quality, dashboard",
        "purpose": "Incidentes operacionais e de DQ.",
        "orm_model": "Incident",
        "orm_file": "backend/app/models/incident.py",
        "migration": "5e0cb6f8c1a2_add_incidents_table_t2c_ops.py; d1a8f9c77b21_add_incident_source_fields_for_dq.py; a7b5fe120b3f_add_incident_source_columns_for_dq_rule_link.py; f2c7b0de91aa_harden_incident_dq_columns_and_status.py",
        "writers": "backend/app/api/incidents.py; backend/app/services/dq.py",
        "readers": "incidents, dashboard, certification, home",
        "trigger_endpoint": "CRUD /api/v1/incidents/*; geração automática em DQ",
        "relationships": "refs opcionais a users; associação lógica com tables por table_fqn",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "users",
        "module": "auth, admin",
        "purpose": "Usuários da plataforma.",
        "orm_model": "User",
        "orm_file": "backend/app/models/auth.py",
        "migration": "55da547ae3ae_baseline.py; c9f8b65af01d_add_user_name_column.py",
        "writers": "backend/app/seed.py; backend/app/api/admin.py; backend/app/api/me.py",
        "readers": "auth, ownership, audit, certification, privacy, admin",
        "trigger_endpoint": "Startup seed dev; CRUD /api/v1/admin/users*; PUT/PATCH /api/v1/me",
        "relationships": "N:M com roles; referenciado por audit, certification, privacy e incidents",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "roles",
        "module": "auth, admin",
        "purpose": "Perfis/Roles do sistema.",
        "orm_model": "Role",
        "orm_file": "backend/app/models/auth.py",
        "migration": "55da547ae3ae_baseline.py",
        "writers": "backend/app/seed.py; backend/app/api/admin.py",
        "readers": "auth, admin",
        "trigger_endpoint": "Startup seed dev; CRUD /api/v1/admin/roles*",
        "relationships": "N:M com users; N:M com permissions",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "permissions",
        "module": "auth, admin",
        "purpose": "Permissões finas de acesso.",
        "orm_model": "Permission",
        "orm_file": "backend/app/models/auth.py",
        "migration": "7c5bf86829b1_add_permissions_rbac.py",
        "writers": "backend/app/seed.py; backend/app/api/admin.py",
        "readers": "auth, admin",
        "trigger_endpoint": "Startup seed dev; CRUD /api/v1/admin/permissions*",
        "relationships": "N:M com roles",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "user_role",
        "module": "auth, admin",
        "purpose": "Associação usuário-role.",
        "orm_model": "Tabela de associação user_role",
        "orm_file": "backend/app/models/auth.py",
        "migration": "55da547ae3ae_baseline.py",
        "writers": "backend/app/seed.py; backend/app/api/admin.py",
        "readers": "auth, admin",
        "trigger_endpoint": "Startup seed dev; CRUD indireto de users/roles",
        "relationships": "N:M users x roles",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "role_permissions",
        "module": "auth, admin",
        "purpose": "Associação role-permission.",
        "orm_model": "Tabela de associação role_permissions",
        "orm_file": "backend/app/models/auth.py",
        "migration": "7c5bf86829b1_add_permissions_rbac.py",
        "writers": "backend/app/seed.py; backend/app/api/admin.py",
        "readers": "auth, admin",
        "trigger_endpoint": "Startup seed dev; CRUD indireto de roles/permissions",
        "relationships": "N:M roles x permissions",
        "status": "ativo",
    },
    {
        "bucket": "Persistência interna",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "audit_log",
        "module": "audit transversal",
        "purpose": "Auditoria operacional das mudanças manuais relevantes.",
        "orm_model": "AuditLog",
        "orm_file": "backend/app/models/audit.py",
        "migration": "2f1c4be8aa10_create_audit_log_table.py",
        "writers": "backend/app/services/audit.py",
        "readers": "backend/app/api/audit.py",
        "trigger_endpoint": "Escrita indireta por datasource, catalog, table_metadata, lineage, certification, privacy_access, incidents, admin e import_export",
        "relationships": "FK opcional para users",
        "status": "ativo",
    },
    {
        "bucket": "Legado / compatibilidade",
        "database": "DATABASE_URL",
        "schema": "t2c_data",
        "table": "audit_logs",
        "module": "audit legado",
        "purpose": "Tabela antiga de auditoria criada na baseline.",
        "orm_model": "Sem model ativo atual",
        "orm_file": "N/A",
        "migration": "55da547ae3ae_baseline.py",
        "writers": "Nenhum writer principal atual encontrado",
        "readers": "Nenhum fluxo principal atual identificado",
        "trigger_endpoint": "N/A",
        "relationships": "Legado",
        "status": "legado",
    },
]


SUMMARY_ROWS = [
    ("Banco interno principal", "PostgreSQL via DATABASE_URL"),
    ("Schema principal", "t2c_data"),
    ("Schema secundário", "t2c_ops"),
    ("Persistência oficial", "Metadados, governança, DQ, lineage, auth/admin, incidentes e auditoria"),
    ("Tabelas externas catalogadas", "Não são copiadas fisicamente; são representadas por data_sources/databases/schemas/tables/columns"),
    ("Módulos sem tabela própria", "dashboard, home/resumo, search, certification e privacy/access usam agregações/campos em tabelas existentes"),
    ("Legados destacados", "audit_logs plural; a linhagem canônica agora usa lineage_assets/relations/jobs/runs/source_configs."),
]


FILES_ANALYZED = [
    "backend/app/core/db.py",
    "backend/app/core/config.py",
    "backend/app/models/catalog.py",
    "backend/app/models/dq.py",
    "backend/app/models/auth.py",
    "backend/app/models/lineage.py",
    "backend/app/models/scan.py",
    "backend/app/models/tag.py",
    "backend/app/models/glossary.py",
    "backend/app/models/audit.py",
    "backend/app/models/incident.py",
    "backend/app/api/router.py",
    "backend/app/api/datasource.py",
    "backend/app/api/scan.py",
    "backend/app/api/catalog.py",
    "backend/app/api/table_metadata.py",
    "backend/app/api/dq.py",
    "backend/app/api/lineage.py",
    "backend/app/api/incidents.py",
    "backend/app/api/data_owners.py",
    "backend/app/api/privacy_access.py",
    "backend/app/api/certification.py",
    "backend/app/api/admin.py",
    "backend/app/api/me.py",
    "backend/app/api/import_export.py",
    "backend/app/services/scanner.py",
    "backend/app/services/dq.py",
    "backend/app/services/dq_spark.py",
    "backend/app/services/lineage.py",
    "backend/app/services/lineage_spreadsheet.py",
    "backend/app/features/lineage/openlineage_sync.py",
    "backend/app/services/dashboard_summary.py",
    "backend/app/services/audit.py",
    "backend/app/seed.py",
    "backend/alembic/versions/55da547ae3ae_baseline.py",
    "backend/alembic/versions/3c1f0b7d2e4a_add_data_owners.py",
    "backend/alembic/versions/a31c5d9f4b22_add_table_certification_fields.py",
    "backend/alembic/versions/c4e7a2b1d9f0_add_certification_badges_column.py",
    "backend/alembic/versions/f7c1d2e3a4b5_add_table_privacy_access_fields.py",
    "backend/alembic/versions/e6d4a1b9c2f3_add_connection_config_to_datasources.py",
    "backend/alembic/versions/e0a4d2c3b901_add_dq_observability_tables.py",
    "backend/alembic/versions/9f2a6d1d77aa_add_dq_rules_tables.py",
    "backend/alembic/versions/b1d3a9c2f401_add_dq_job_runs_for_spark_execution.py",
    "backend/alembic/versions/1d6c80ac7e58_add_spark_fields_to_dq_runs.py",
    "backend/alembic/versions/85e30b8d49cd_add_execution_engine_to_dq_runs.py",
    "backend/alembic/versions/6c0464c9c46a_add_dq_run_id_to_dq_job_runs.py",
    "backend/alembic/versions/7c5bf86829b1_add_permissions_rbac.py",
    "backend/alembic/versions/c9f8b65af01d_add_user_name_column.py",
    "backend/alembic/versions/2f1c4be8aa10_create_audit_log_table.py",
    "backend/alembic/versions/5e0cb6f8c1a2_add_incidents_table_t2c_ops.py",
    "backend/alembic/versions/d1a8f9c77b21_add_incident_source_fields_for_dq.py",
    "backend/alembic/versions/a7b5fe120b3f_add_incident_source_columns_for_dq_rule_link.py",
    "backend/alembic/versions/f2c7b0de91aa_harden_incident_dq_columns_and_status.py",
    "backend/alembic/versions/ab39c8d4e2f1_add_manual_lineage_assets_and_relations.py",
    "backend/alembic/versions/0b1c2d3e4f5a_add_internal_openlineage_tables.py",
    "backend/alembic/versions/9b7e3c2d1f44_add_column_dictionary_fields.py",
]


NOTES = [
    ("Nome do banco", "O banco físico varia por ambiente e vem de DATABASE_URL. No ambiente Docker observado, aparece como andromeda."),
    ("Certification e Privacy", "Não possuem tabela própria; persistem em campos adicionais de t2c_data.tables."),
    ("Dashboard/Home/Search", "São módulos majoritariamente de leitura/agregação sobre tabelas já persistidas."),
    ("Audit", "Há coexistência entre audit_logs (baseline) e audit_log (modelo/uso atual)."),
    ("Lineage", "Há coexistência entre lineage legado e lineage atual. O fluxo principal atual usa lineage_assets/lineage_relations."),
    ("Repository layer", "O backend atual não padroniza repositories; a gravação ocorre em services e, em alguns casos, direto nas APIs FastAPI."),
]


def autosize(ws) -> None:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 60)


def style_header(row) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    fill = PatternFill("solid", fgColor="E2E8F0")
    font = Font(bold=True, color="0F172A")
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Resumo Executivo"
    ws_summary.append(["Item", "Valor"])
    style_header(ws_summary[1])
    for row in SUMMARY_ROWS:
        ws_summary.append(list(row))
    autosize(ws_summary)

    ws_dict = wb.create_sheet("Dicionario Consolidado")
    headers = [
        "Categoria",
        "Banco",
        "Schema",
        "Tabela",
        "Modulo",
        "Finalidade",
        "Model ORM",
        "Arquivo ORM",
        "Migration",
        "Quem grava",
        "Quem le",
        "Endpoint que aciona gravacao",
        "Relacionamentos",
        "Status",
    ]
    ws_dict.append(headers)
    style_header(ws_dict[1])
    for row in ROWS:
        ws_dict.append(
            [
                row["bucket"],
                row["database"],
                row["schema"],
                row["table"],
                row["module"],
                row["purpose"],
                row["orm_model"],
                row["orm_file"],
                row["migration"],
                row["writers"],
                row["readers"],
                row["trigger_endpoint"],
                row["relationships"],
                row["status"],
            ]
        )
    for row in ws_dict.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    table = Table(displayName="DicionarioPersistencia", ref=f"A1:N{ws_dict.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws_dict.add_table(table)
    autosize(ws_dict)

    ws_files = wb.create_sheet("Arquivos Analisados")
    ws_files.append(["Arquivo"])
    style_header(ws_files[1])
    for item in FILES_ANALYZED:
        ws_files.append([item])
    autosize(ws_files)

    ws_notes = wb.create_sheet("Duvidas e Inferencias")
    ws_notes.append(["Tema", "Observacao"])
    style_header(ws_notes[1])
    for note in NOTES:
        ws_notes.append(list(note))
    autosize(ws_notes)

    ws_cover = wb.create_sheet("Legenda")
    ws_cover.append(["Chave", "Descricao"])
    style_header(ws_cover[1])
    legend_rows = [
        ("Persistência interna", "Tabelas oficiais usadas pelo t2c_data para armazenar estado do produto."),
        ("Legado / compatibilidade", "Estruturas herdadas que ainda coexistem, mas não são o caminho principal."),
        ("DATABASE_URL", "Nome do banco físico vem da configuração de ambiente, não do código."),
        ("Quem grava", "Service ou API FastAPI que persiste diretamente os registros."),
    ]
    for row in legend_rows:
        ws_cover.append(list(row))
    autosize(ws_cover)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row == 1:
                    continue
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False

    accent_fill = PatternFill("solid", fgColor="F8FAFC")
    for title_cell in (ws_summary["A1"], ws_dict["A1"], ws_files["A1"], ws_notes["A1"], ws_cover["A1"]):
        title_cell.fill = accent_fill
        title_cell.font = Font(bold=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    if len(sys.argv) > 1:
        output_path = Path(sys.argv[1]).resolve()
    else:
        repo_root = Path(__file__).resolve().parents[2]
        output_path = repo_root / "docs" / "dicionario_dados_persistencia_t2c_data.xlsx"
    build_workbook(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
