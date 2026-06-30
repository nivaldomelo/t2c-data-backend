from __future__ import annotations

import os
import importlib.util
from io import BytesIO

os.environ.setdefault("DATABASE_URL", "sqlite+pysqlite:///:memory:")
os.environ.setdefault("ENV", "test")

from sqlalchemy import create_engine, event, select
from sqlalchemy.dialects.sqlite.base import SQLiteTypeCompiler
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from t2c_data.models import Base
from t2c_data.models.catalog import ColumnEntity, DataSource, Database, Schema, TableEntity
from t2c_data.features.catalog.column_dictionary_admin import (
    ColumnDictionaryFilters,
    get_column_dictionary_summary,
    list_column_dictionary,
    reset_column_dictionary_curation,
    preview_column_dictionary_import,
    update_column_dictionary_item,
)
from t2c_data.features.catalog.column_dictionary_workbook import build_column_dictionary_workbook, import_column_dictionary_from_workbook
from t2c_data.schemas.column_dictionary import ColumnDictionaryUpdateIn


OPENPYXL_AVAILABLE = importlib.util.find_spec("openpyxl") is not None


if not hasattr(SQLiteTypeCompiler, "visit_INET"):
    SQLiteTypeCompiler.visit_INET = lambda self, type_, **kw: "TEXT"  # type: ignore[attr-defined]
if not hasattr(SQLiteTypeCompiler, "visit_JSONB"):
    SQLiteTypeCompiler.visit_JSONB = lambda self, type_, **kw: "JSON"  # type: ignore[attr-defined]


def _build_session() -> Session:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        future=True,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine, "connect")
    def _attach_schema(dbapi_connection, _connection_record) -> None:  # type: ignore[no-untyped-def]
        cursor = dbapi_connection.cursor()
        cursor.execute("ATTACH DATABASE ':memory:' AS t2c_data")
        cursor.execute("ATTACH DATABASE ':memory:' AS t2c_ops")
        cursor.close()

    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, class_=Session)
    return SessionLocal()


def _seed_catalog(db: Session) -> tuple[ColumnEntity, ColumnEntity, ColumnEntity]:
    datasource = DataSource(
        name="postgres-primary",
        db_type="postgres",
        host="localhost",
        port=5432,
        database="andromeda",
        username="nivasmelo",
    )
    datasource.password = "secret"
    database = Database(name="andromeda", datasource=datasource)
    schema_public = Schema(name="public", database=database)
    schema_finance = Schema(name="finance", database=database)
    table_customers = TableEntity(name="customers", table_type="table", schema=schema_public)
    table_orders = TableEntity(name="orders", table_type="table", schema=schema_finance)
    table_customers.description_source = "Tabela de clientes"
    table_orders.description_source = "Tabela de pedidos"
    db.add_all([datasource, database, schema_public, schema_finance, table_customers, table_orders])
    db.flush()

    column_id = ColumnEntity(
        table=table_customers,
        name="id",
        data_type="integer",
        is_primary_key=True,
        is_nullable=False,
        ordinal_position=1,
        description_source="Identificador do cliente",
        dictionary_description="Identificador único do cliente",
        dictionary_comment="Chave principal da entidade cliente",
        existing_comment="Cliente PK",
        slug="public-customers-id",
        external_id="col-001",
    )
    column_name = ColumnEntity(
        table=table_customers,
        name="name",
        data_type="varchar",
        is_primary_key=False,
        is_nullable=False,
        ordinal_position=2,
        description_source="Nome do cliente",
    )
    column_amount = ColumnEntity(
        table=table_orders,
        name="amount",
        data_type="numeric",
        is_primary_key=False,
        is_nullable=False,
        ordinal_position=1,
        description_source=None,
        existing_comment=None,
    )
    db.add_all([column_id, column_name, column_amount])
    db.commit()
    return column_id, column_name, column_amount


def test_column_dictionary_summary_and_listing() -> None:
    db = _build_session()
    column_id, column_name, column_amount = _seed_catalog(db)

    summary = get_column_dictionary_summary(db, filters=ColumnDictionaryFilters())
    page = list_column_dictionary(db, filters=ColumnDictionaryFilters(schema_name="public"), page=1, page_size=10)

    assert summary.total_columns == 3
    assert summary.total_tables == 2
    assert summary.total_schemas == 2
    assert summary.documented_columns == 1
    assert summary.comment_columns == 1
    assert summary.existing_comment_columns == 1
    assert summary.pending_columns == 2
    assert summary.top_gap_tables[0].schema_name == "finance"
    assert summary.top_gap_tables[0].table_name == "orders"

    assert page.total == 2
    assert [item.name for item in page.items] == ["id", "name"]
    assert page.filters.schemas == ["finance", "public"]
    assert page.filters.data_types == ["integer", "numeric", "varchar"]


def test_preview_import_matches_exported_workbook_and_update_applies() -> None:
    if not OPENPYXL_AVAILABLE:
        return

    db = _build_session()
    column_id, column_name, column_amount = _seed_catalog(db)

    column_name.slug = "public-customers-name"
    column_name.external_id = "col-002"
    column_name.dictionary_description = "Nome do cliente"
    column_name.dictionary_comment = "Identificação nominal do cliente"
    column_amount.slug = "finance-orders-amount"
    column_amount.external_id = "col-003"
    column_amount.dictionary_description = "Valor do pedido"
    column_amount.dictionary_comment = "Valor monetário consolidado"
    db.commit()

    workbook = build_column_dictionary_workbook(
        [
            (column_id, "public", "customers"),
            (column_name, "public", "customers"),
            (column_amount, "finance", "orders"),
        ],
        include_readme=True,
    )
    preview = preview_column_dictionary_import(db, workbook)

    assert preview.processed == 3
    assert preview.ignored == 3
    assert preview.inserted == 0
    assert preview.updated == 0
    assert preview.rejected == 0

    updated = update_column_dictionary_item(
        db,
        column_id=column_name.id,
        payload=ColumnDictionaryUpdateIn(
            dictionary_description="Descrição do nome do cliente",
            dictionary_comment="Comentário operacional",
            existing_comment="Comentário original",
        ),
    )
    assert updated.dictionary_description == "Descrição do nome do cliente"
    assert updated.dictionary_comment == "Comentário operacional"
    assert updated.existing_comment == "Comentário original"


def test_import_accepts_header_variants_and_updates_when_sheet_is_not_active() -> None:
    if not OPENPYXL_AVAILABLE:
        return

    from openpyxl import load_workbook

    db = _build_session()
    column_id, column_name, column_amount = _seed_catalog(db)

    column_name.slug = "public-customers-name"
    column_name.external_id = "col-002"
    column_name.dictionary_description = "Nome original"
    column_name.dictionary_comment = "Comentário original"
    db.commit()

    workbook_bytes = build_column_dictionary_workbook(
        [
            (column_id, "public", "customers"),
            (column_name, "public", "customers"),
            (column_amount, "finance", "orders"),
        ],
        include_readme=True,
    )

    workbook = load_workbook(BytesIO(workbook_bytes))
    workbook.create_sheet("Rascunho", 0)
    workbook.active = 0
    sheet = workbook["Colunas_Importacao"]
    sheet["A3"] = "col-002"
    sheet["P1"] = "Descrição"
    sheet["Q1"] = "Comentário"
    sheet["E1"] = "Posição Coluna"
    sheet["P3"] = "Descrição do nome do cliente"
    sheet["Q3"] = "Comentário do nome do cliente"
    sheet["P4"] = "Descrição dos pedidos"
    sheet["Q4"] = "Comentário dos pedidos"

    buffer = BytesIO()
    workbook.save(buffer)

    preview = preview_column_dictionary_import(db, buffer.getvalue())
    result = import_column_dictionary_from_workbook(db, buffer.getvalue())

    assert preview.processed == 3
    assert preview.matched == 3
    assert preview.ignored == 1
    assert preview.rejected == 0
    assert result.processed == 3
    assert result.matched == 3
    assert result.ignored == 1
    assert result.updated == 2
    assert result.rejected == 0

    refreshed_name = db.get(ColumnEntity, column_name.id)
    refreshed_amount = db.get(ColumnEntity, column_amount.id)
    assert refreshed_name is not None
    assert refreshed_name.dictionary_description == "Descrição do nome do cliente"
    assert refreshed_name.dictionary_comment == "Comentário do nome do cliente"
    assert refreshed_amount is not None
    assert refreshed_amount.dictionary_description == "Descrição dos pedidos"
    assert refreshed_amount.dictionary_comment == "Comentário dos pedidos"


def test_import_recreates_rows_after_dictionary_reset() -> None:
    if not OPENPYXL_AVAILABLE:
        return

    db = _build_session()
    _, column_name, column_amount = _seed_catalog(db)

    workbook = build_column_dictionary_workbook(
        [
            (column_name, "public", "customers"),
            (column_amount, "finance", "orders"),
        ],
        include_readme=True,
    )

    deleted = reset_column_dictionary_curation(db)
    assert deleted == 3
    assert db.scalar(select(ColumnEntity.id)) is None

    preview = preview_column_dictionary_import(db, workbook)
    result = import_column_dictionary_from_workbook(db, workbook)

    assert preview.processed == 2
    assert preview.matched == 2
    assert preview.inserted == 2
    assert preview.updated == 0
    assert preview.rejected == 0

    assert result.processed == 2
    assert result.matched == 2
    assert result.imported == 2
    assert result.updated == 0
    assert result.ignored == 0
    assert result.rejected == 0

    recreated_columns = db.scalars(select(ColumnEntity)).all()
    assert len(recreated_columns) == 2
    assert {column.name for column in recreated_columns} == {"name", "amount"}


if __name__ == "__main__":
    test_column_dictionary_summary_and_listing()
    test_preview_import_matches_exported_workbook_and_update_applies()
    print("column dictionary admin tests: OK")
