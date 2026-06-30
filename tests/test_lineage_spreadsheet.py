from __future__ import annotations

import os
from io import BytesIO

os.environ.setdefault("DATABASE_URL", "sqlite+pysqlite:///:memory:")
os.environ.setdefault("ENV", "test")

from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from t2c_data.core.config import settings
from t2c_data.models.lineage import LineageAsset, LineageRelation
from t2c_data.services.lineage import list_relations_out
from t2c_data.services.lineage_spreadsheet import commit_lineage_import, preview_lineage_import


def _session_factory():
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True).execution_options(
        schema_translate_map={settings.db_schema: None}
    )
    with engine.begin() as conn:
        conn.exec_driver_sql("CREATE TABLE users (id INTEGER PRIMARY KEY)")
        conn.exec_driver_sql("CREATE TABLE data_sources (id INTEGER PRIMARY KEY)")
        conn.exec_driver_sql("CREATE TABLE tables (id INTEGER PRIMARY KEY)")
        LineageAsset.__table__.create(bind=conn)
        LineageRelation.__table__.create(bind=conn)
    return sessionmaker(bind=engine, future=True)


def _workbook_bytes(
    *,
    assets: list[list[object | None]],
    mappings: list[list[object | None]] | None = None,
    relations: list[list[object | None]] | None = None,
) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "1_INSTRUCOES"
    ws.append(["Aba", "Objetivo"])

    ativos = workbook.create_sheet("2_ATIVOS")
    ativos.append(
        [
            "asset_key",
            "camada",
            "tipo_ativo",
            "sistema",
            "schema_name",
            "object_name",
            "nome_exibicao",
            "owner",
            "certificado",
            "ativo",
            "descricao",
            "observacoes",
        ]
    )
    for row in assets:
        ativos.append(row)

    mapping = workbook.create_sheet("3_MAPEAMENTO_NEGOCIO")
    mapping.append(
        [
            "camada",
            "tipo_ativo",
            "schema_name",
            "object_name",
            "asset_key",
            "origem_externa_tipo",
            "origem_externa_nome",
            "upstream_asset_keys",
            "process_name",
            "process_type",
            "dashboards_consumidores",
            "observacoes",
        ]
    )
    for row in mappings or []:
        mapping.append(row)

    rels = workbook.create_sheet("4_LINHAGEM_IMPORTACAO")
    rels.append(
        [
            "source_asset_key",
            "target_asset_key",
            "relation_type",
            "process_name",
            "process_type",
            "notes",
            "ativo",
        ]
    )
    for row in relations or []:
        rels.append(row)

    lists = workbook.create_sheet("5_LISTAS")
    lists.append(["Lista", "Valores"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _base_assets() -> list[list[object | None]]:
    return [
        ["table.bronze.proposta_raw", "bronze", "table", "demo", "bronze", "proposta_raw", "bronze.proposta_raw", None, "nao", "sim", None, None],
        ["table.silver.proposta_tratada", "silver", "table", "demo", "silver", "proposta_tratada", "silver.proposta_tratada", None, "nao", "sim", None, None],
        ["table.gold.fato_vendas", "gold", "table", "demo", "gold", "fato_vendas", "gold.fato_vendas", None, "nao", "sim", None, None],
    ]


def test_import_only_assets_warns_absence_of_relations():
    SessionLocal = _session_factory()
    content = _workbook_bytes(assets=_base_assets())

    with SessionLocal() as session:
        preview = preview_lineage_import(session, content)
        result = commit_lineage_import(session, content)
        assets = session.scalars(select(LineageAsset)).all()
        relations = session.scalars(select(LineageRelation)).all()

    assert preview["summary"]["edges_found"] == 0
    assert result["edges_found"] == 0
    assert len(assets) == 3
    assert len(relations) == 0
    assert any("no lineage relations were generated" in item["message"] for item in result["warnings"])


def test_import_mapping_generates_upstream_relations():
    SessionLocal = _session_factory()
    content = _workbook_bytes(
        assets=_base_assets(),
        mappings=[
            ["silver", "table", "silver", "proposta_tratada", "table.silver.proposta_tratada", None, None, "table.bronze.proposta_raw", "job_silver", "airflow", None, "tratamento"],
            ["gold", "table", "gold", "fato_vendas", "table.gold.fato_vendas", None, None, "table.silver.proposta_tratada", "job_gold", "airflow", None, "gold"],
        ],
    )

    with SessionLocal() as session:
        result = commit_lineage_import(session, content)
        relations = session.scalars(select(LineageRelation).order_by(LineageRelation.id)).all()
        relation_keys = [(item.source_asset.asset_key, item.target_asset.asset_key, item.relation_type) for item in relations]

    assert result["edges_found"] == 2
    assert result["created_relations"] == 2
    assert relation_keys == [
        ("table.bronze.proposta_raw", "table.silver.proposta_tratada", "transformation"),
        ("table.silver.proposta_tratada", "table.gold.fato_vendas", "transformation"),
    ]


def test_import_mapping_generates_dashboard_consumption_relations():
    SessionLocal = _session_factory()
    content = _workbook_bytes(
        assets=_base_assets(),
        mappings=[
            ["gold", "table", "gold", "fato_vendas", "table.gold.fato_vendas", None, None, None, "job_gold", "airflow", "comercial_kpis; financeiro_kpis", "dashboards"],
        ],
    )

    with SessionLocal() as session:
        result = commit_lineage_import(session, content)
        relations = session.scalars(select(LineageRelation).order_by(LineageRelation.id)).all()
        dashboards = session.scalars(select(LineageAsset).where(LineageAsset.asset_type == "dashboard")).all()

    assert result["edges_found"] == 2
    assert result["created_relations"] == 2
    assert len(dashboards) == 2
    assert all(item.relation_type == "consumption" for item in relations)


def test_import_explicit_sheet_generates_edges():
    SessionLocal = _session_factory()
    content = _workbook_bytes(
        assets=_base_assets(),
        relations=[
            ["table.silver.proposta_tratada", "table.gold.fato_vendas", "load", "job_gold", "dbt", "edge explicita", "sim"],
        ],
    )

    with SessionLocal() as session:
        result = commit_lineage_import(session, content)
        relations = session.scalars(select(LineageRelation)).all()
        relation_type = relations[0].relation_type
        relation_notes = relations[0].notes

    assert result["edges_found"] == 1
    assert result["created_relations"] == 1
    assert relation_type == "load"
    assert relation_notes == "edge explicita"


def test_lineage_listing_reflects_imported_relations():
    SessionLocal = _session_factory()
    content = _workbook_bytes(
        assets=_base_assets(),
        relations=[
            ["table.silver.proposta_tratada", "table.gold.fato_vendas", "load", "job_gold", "dbt", "edge explicita", "sim"],
        ],
    )

    with SessionLocal() as session:
        commit_lineage_import(session, content)
        listing = list_relations_out(session)

    assert listing.summary.total_assets == 3
    assert listing.summary.total_relations == 1
    assert len(listing.items) == 1
    assert listing.items[0].source_asset.asset_key == "table.silver.proposta_tratada"
    assert listing.items[0].target_asset.asset_key == "table.gold.fato_vendas"
