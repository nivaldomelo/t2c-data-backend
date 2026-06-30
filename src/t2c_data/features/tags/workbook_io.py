from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from io import BytesIO
import re
import unicodedata

from sqlalchemy import select
from sqlalchemy.orm import Session

from t2c_data.features.export_security import redact_export_value
from t2c_data.models.tag import Tag
from t2c_data.schemas.tag import TagSpreadsheetImportError, TagSpreadsheetImportResult

TAG_SPREADSHEET_HEADERS = [
    "ID",
    "Slug",
    "Tag",
    "Grupo",
    "Subgrupo",
    "Descricao",
    "Exemplo_de_Uso",
    "Tipo_de_Tag",
    "Escopo_Sugerido",
    "Status",
    "Sinonimos",
    "Observacoes",
]

ALLOWED_TAG_STATUS = {
    "ativo": "active",
    "active": "active",
    "inativo": "inactive",
    "inactive": "inactive",
    "rascunho": "draft",
    "draft": "draft",
    "deprecated": "deprecated",
    "descontinuado": "deprecated",
    "archived": "archived",
    "arquivado": "archived",
}

STATUS_LABELS = {
    "active": "Ativo",
    "inactive": "Inativo",
    "draft": "Rascunho",
    "deprecated": "Descontinuado",
    "archived": "Arquivado",
}


class TagSpreadsheetError(ValueError):
    pass


@dataclass
class ParsedTagRow:
    row_number: int
    external_id: str | None
    slug: str
    name: str
    group_name: str | None
    subgroup_name: str | None
    description: str | None
    example_of_use: str | None
    tag_type: str | None
    suggested_scope: str | None
    status: str
    synonyms: str | None
    notes: str | None


def slugify_tag(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = normalized.lower().strip()
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized)
    normalized = re.sub(r"-{2,}", "-", normalized)
    return normalized.strip("-")


def normalize_tag_status(value: str | None) -> str:
    raw = (value or "active").strip().lower()
    if not raw:
        return "active"
    normalized = ALLOWED_TAG_STATUS.get(raw)
    if normalized is None:
        raise TagSpreadsheetError(
            f"Status inválido: {value}. Use um dos valores: {', '.join(sorted(STATUS_LABELS.values()))}."
        )
    return normalized


def status_label(value: str | None) -> str:
    normalized = normalize_tag_status(value)
    return STATUS_LABELS.get(normalized, normalized.title())


def _normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _iter_rows_from_workbook(content: bytes) -> Iterable[tuple[int, list[object | None]]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - depends on runtime deps
        raise TagSpreadsheetError("Dependência openpyxl não instalada no backend.") from exc

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise TagSpreadsheetError("Arquivo inválido. Envie uma planilha .xlsx compatível com o modelo.") from exc

    sheet = workbook["Tags_Importacao"] if "Tags_Importacao" in workbook.sheetnames else workbook.active
    header = [
        str(value).strip() if value is not None else ""
        for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    if header[: len(TAG_SPREADSHEET_HEADERS)] != TAG_SPREADSHEET_HEADERS:
        raise TagSpreadsheetError(
            "Cabeçalho inválido. Utilize o modelo com as colunas: " + ", ".join(TAG_SPREADSHEET_HEADERS)
        )

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[: len(TAG_SPREADSHEET_HEADERS)])
        if not any(value not in (None, "") for value in values):
            continue
        yield index, values


def parse_tag_workbook(content: bytes) -> tuple[list[ParsedTagRow], list[TagSpreadsheetImportError]]:
    parsed: list[ParsedTagRow] = []
    errors: list[TagSpreadsheetImportError] = []

    for row_number, values in _iter_rows_from_workbook(content):
        external_id = _normalize_text(values[0])
        slug = _normalize_text(values[1])
        name = _normalize_text(values[2])
        group_name = _normalize_text(values[3])
        subgroup_name = _normalize_text(values[4])
        description = _normalize_text(values[5])
        example_of_use = _normalize_text(values[6])
        tag_type = _normalize_text(values[7])
        suggested_scope = _normalize_text(values[8])
        status_value = _normalize_text(values[9])
        synonyms = _normalize_text(values[10])
        notes = _normalize_text(values[11])

        if not name:
            errors.append(TagSpreadsheetImportError(row_number=row_number, slug=slug, message="Tag é obrigatória."))
            continue

        final_slug = slugify_tag(slug or name)
        if not final_slug:
            errors.append(TagSpreadsheetImportError(row_number=row_number, slug=slug, message="Slug inválido."))
            continue

        try:
            normalized_status = normalize_tag_status(status_value)
        except TagSpreadsheetError as exc:
            errors.append(TagSpreadsheetImportError(row_number=row_number, slug=final_slug, message=str(exc)))
            continue

        parsed.append(
            ParsedTagRow(
                row_number=row_number,
                external_id=external_id,
                slug=final_slug,
                name=name,
                group_name=group_name,
                subgroup_name=subgroup_name,
                description=description,
                example_of_use=example_of_use,
                tag_type=tag_type,
                suggested_scope=suggested_scope,
                status=normalized_status,
                synonyms=synonyms,
                notes=notes,
            )
        )

    return parsed, errors


def import_tags_from_workbook(session: Session, content: bytes) -> TagSpreadsheetImportResult:
    parsed_rows, parsing_errors = parse_tag_workbook(content)
    errors = list(parsing_errors)
    imported = 0
    updated = 0
    rejected = len(errors)

    seen_slugs: set[str] = set()
    seen_external_ids: set[str] = set()
    for row in parsed_rows:
        if row.slug in seen_slugs:
            errors.append(TagSpreadsheetImportError(row_number=row.row_number, slug=row.slug, message="Slug duplicado na planilha."))
            rejected += 1
            continue
        seen_slugs.add(row.slug)

        if row.external_id:
            if row.external_id in seen_external_ids:
                errors.append(
                    TagSpreadsheetImportError(
                        row_number=row.row_number,
                        slug=row.slug,
                        message="ID duplicado na planilha.",
                    )
                )
                rejected += 1
                continue
            seen_external_ids.add(row.external_id)

        tag = session.scalar(select(Tag).where(Tag.slug == row.slug))
        if tag is None and row.external_id:
            tag = session.scalar(select(Tag).where(Tag.external_id == row.external_id))

        if tag is None:
            tag = Tag(
                external_id=row.external_id,
                slug=row.slug,
                name=row.name,
                group_name=row.group_name,
                subgroup_name=row.subgroup_name,
                description=row.description,
                example_of_use=row.example_of_use,
                tag_type=row.tag_type,
                suggested_scope=row.suggested_scope,
                status=row.status,
                synonyms=row.synonyms,
                notes=row.notes,
            )
            session.add(tag)
            imported += 1
            continue

        tag.external_id = row.external_id or tag.external_id
        tag.slug = row.slug
        tag.name = row.name
        tag.group_name = row.group_name
        tag.subgroup_name = row.subgroup_name
        tag.description = row.description
        tag.example_of_use = row.example_of_use
        tag.tag_type = row.tag_type
        tag.suggested_scope = row.suggested_scope
        tag.status = row.status
        tag.synonyms = row.synonyms
        tag.notes = row.notes
        updated += 1

    session.commit()
    return TagSpreadsheetImportResult(
        processed=len(parsed_rows) + len(parsing_errors),
        imported=imported,
        updated=updated,
        rejected=rejected,
        errors=errors,
    )


def tag_export_rows(tags: Iterable[Tag]) -> list[list[str]]:
    rows: list[list[str]] = []
    for tag in tags:
        rows.append(
            [
                redact_export_value(tag.external_id, field_name="external_id"),
                redact_export_value(tag.slug, field_name="slug"),
                redact_export_value(tag.name, field_name="name"),
                redact_export_value(tag.group_name, field_name="group_name"),
                redact_export_value(tag.subgroup_name, field_name="subgroup_name"),
                redact_export_value(tag.description, field_name="description"),
                redact_export_value(tag.example_of_use, field_name="example_of_use"),
                redact_export_value(tag.tag_type, field_name="tag_type"),
                redact_export_value(tag.suggested_scope, field_name="suggested_scope"),
                redact_export_value(status_label(tag.status), field_name="status"),
                redact_export_value(tag.synonyms, field_name="synonyms"),
                redact_export_value(tag.notes, field_name="notes"),
            ]
        )
    return rows


def build_tag_workbook(tags: Iterable[Tag], *, include_readme: bool = True) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ModuleNotFoundError as exc:  # pragma: no cover - depends on runtime deps
        raise TagSpreadsheetError("Dependência openpyxl não instalada no backend.") from exc

    tag_list = list(tags)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tags_Importacao"
    sheet.append(TAG_SPREADSHEET_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="F97316")
    for row in tag_export_rows(tag_list):
        sheet.append(row)

    if sheet.max_row > 1:
        table = Table(displayName="tblTags", ref=f"A1:L{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)

    widths = [14, 24, 26, 18, 18, 52, 42, 20, 24, 16, 24, 28]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    if include_readme:
        readme = workbook.create_sheet("README")
        readme.append(["Campo", "Descrição"])
        for cell in readme[1]:
            cell.font = Font(bold=True)
        instructions = [
            ("ID", "Identificador externo opcional para importações futuras."),
            ("Slug", "Chave técnica estável e única. É a referência principal da carga."),
            ("Tag", "Nome visível da tag."),
            ("Grupo / Subgrupo", "Taxonomia para navegação e filtros."),
            ("Status", "Aceita Ativo, Inativo, Rascunho, Descontinuado e Arquivado."),
        ]
        for row in instructions:
            readme.append(row)
        summary = workbook.create_sheet("Resumo")
        summary.append(["Indicador", "Valor"])
        summary.append(["Total de tags", len(tag_list)])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
