from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import re

LINEAGE_ASSET_HEADERS = [
    "asset_key",
    "camada",
    "tipo_ativo",
    "sistema",
    "schema_name",
    "object_name",
    "nome_exibicao",
    "owner",
    "certificado",
    "ativo",
    "descricao",
    "observacoes",
]

LINEAGE_MAPPING_HEADERS = [
    "camada",
    "tipo_ativo",
    "schema_name",
    "object_name",
    "asset_key",
    "origem_externa_tipo",
    "origem_externa_nome",
    "upstream_asset_keys",
    "process_name",
    "process_type",
    "dashboards_consumidores",
    "observacoes",
]

LINEAGE_IMPORT_HEADERS = [
    "source_asset_key",
    "target_asset_key",
    "relation_type",
    "process_name",
    "process_type",
    "notes",
    "ativo",
]

SPLIT_PATTERN = re.compile(r"[;,|\n]+")


class LineageSpreadsheetError(ValueError):
    pass


@dataclass
class ParsedAssetRow:
    row_number: int
    asset_key: str
    layer: str
    asset_type: str
    system_name: str | None
    schema_name: str | None
    object_name: str | None
    display_name: str
    owner: str | None
    certified: bool | None
    is_active: bool
    description: str | None
    notes: str | None


@dataclass
class ParsedMappingRow:
    row_number: int
    layer: str | None
    asset_type: str | None
    schema_name: str | None
    object_name: str | None
    asset_key: str
    source_kind: str | None
    source_name: str | None
    upstream_asset_keys: list[str]
    process_name: str | None
    process_type: str | None
    dashboards: list[str]
    notes: str | None


@dataclass
class ParsedRelationRow:
    row_number: int
    source_asset_key: str
    target_asset_key: str
    relation_type: str
    process_name: str | None
    process_type: str | None
    notes: str | None
    is_active: bool


def no_relations_warning() -> dict[str, object]:
    return {
        "sheet": "workbook",
        "row_number": 0,
        "message": "Import completed with warnings: assets were imported, but no lineage relations were generated. Fill `4_LINHAGEM_IMPORTACAO` or `upstream_asset_keys` in `3_MAPEAMENTO_NEGOCIO`.",
    }


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_bool(value: object, default: bool = True) -> bool:
    if value in (None, ""):
        return default
    text = str(value).strip().lower()
    if text in {"1", "sim", "s", "ativo", "true", "yes", "y", "x"}:
        return True
    if text in {"0", "nao", "não", "n", "inativo", "false", "no"}:
        return False
    return default


def normalize_asset_type(value: str | None) -> str:
    raw = (value or "table").strip().lower()
    mapping = {
        "table": "table",
        "tabela": "table",
        "view": "view",
        "dashboard": "dashboard",
        "dash": "dashboard",
        "source": "source",
        "origem": "source",
    }
    return mapping.get(raw, raw or "table")


def normalize_layer(value: str | None) -> str:
    raw = (value or "definir").strip().lower()
    mapping = {
        "bronze": "bronze",
        "silver": "silver",
        "gold": "gold",
        "mart": "mart",
        "dashboard": "dashboard",
        "source": "source",
        "origem": "source",
        "definir": "definir",
    }
    return mapping.get(raw, raw or "definir")


def normalize_relation_type(value: str | None) -> str:
    raw = (value or "transformation").strip().lower()
    mapping = {
        "ingestion": "ingestion",
        "transform": "transformation",
        "transformation": "transformation",
        "load": "load",
        "consumption": "consumption",
        "consumo": "consumption",
    }
    return mapping.get(raw, raw)


def split_multi_value(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in SPLIT_PATTERN.split(value) if item.strip()]


def external_source_asset_key(kind: str, name: str) -> str:
    safe_kind = re.sub(r"[^a-z0-9]+", "_", kind.strip().lower()).strip("_") or "other"
    safe_name = re.sub(r"[^a-z0-9]+", "_", name.strip().lower()).strip("_") or "source"
    return f"source.{safe_kind}.{safe_name}"


def display_name_from_parts(schema_name: str | None, object_name: str | None, fallback: str) -> str:
    if schema_name and object_name:
        return f"{schema_name}.{object_name}"
    return object_name or fallback


def excel_table_name(sheet_name: str) -> str:
    sanitized = re.sub(r"[^a-zA-Z0-9_]+", "_", sheet_name).strip("_") or "sheet"
    if sanitized[0].isdigit():
        sanitized = f"tbl_{sanitized}"
    elif not sanitized.startswith("tbl_"):
        sanitized = f"tbl_{sanitized}"
    return sanitized[:255]


def iter_rows_from_workbook(content: bytes, sheet_name: str, headers: list[str]) -> list[tuple[int, list[object | None]]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise LineageSpreadsheetError("Dependência openpyxl não instalada no backend.") from exc

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise LineageSpreadsheetError("Arquivo inválido. Envie uma planilha .xlsx compatível com o modelo.") from exc

    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    header = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if header[: len(headers)] != headers:
        raise LineageSpreadsheetError(f"Cabeçalho inválido na aba {sheet_name}.")

    rows: list[tuple[int, list[object | None]]] = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[: len(headers)])
        if not any(value not in (None, "") for value in values):
            continue
        rows.append((index, values))
    return rows


__all__ = [
    "LINEAGE_ASSET_HEADERS",
    "LINEAGE_IMPORT_HEADERS",
    "LINEAGE_MAPPING_HEADERS",
    "LineageSpreadsheetError",
    "ParsedAssetRow",
    "ParsedMappingRow",
    "ParsedRelationRow",
    "display_name_from_parts",
    "excel_table_name",
    "external_source_asset_key",
    "iter_rows_from_workbook",
    "no_relations_warning",
    "normalize_asset_type",
    "normalize_bool",
    "normalize_layer",
    "normalize_relation_type",
    "normalize_text",
    "split_multi_value",
]
