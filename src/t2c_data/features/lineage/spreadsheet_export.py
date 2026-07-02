from __future__ import annotations

from collections import defaultdict
from io import BytesIO

from sqlalchemy import select
from sqlalchemy.orm import Session

from t2c_data.features.export_security import DEFAULT_EXPORT_LIMIT, enforce_export_limit, redact_export_value, safe_sheet_append
from t2c_data.features.lineage.spreadsheet_parser_support import (
    LINEAGE_ASSET_HEADERS,
    LINEAGE_IMPORT_HEADERS,
    LINEAGE_MAPPING_HEADERS,
    LineageSpreadsheetError,
    excel_table_name,
)
from t2c_data.models.lineage import LineageAsset, LineageRelation


def build_lineage_workbook(session: Session, *, limit: int = DEFAULT_EXPORT_LIMIT) -> tuple[bytes, dict[str, int | bool]]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise LineageSpreadsheetError("Dependência openpyxl não instalada no backend.") from exc

    workbook = Workbook()

    instruction_sheet = workbook.active
    instruction_sheet.title = "1_INSTRUCOES"
    instruction_sheet.append(["Aba", "Objetivo"])
    instruction_sheet.append(["2_ATIVOS", "Cadastre ou revise os ativos participantes da linhagem."])
    instruction_sheet.append(["3_MAPEAMENTO_NEGOCIO", "Mapeie upstreams, origens externas, processos e dashboards consumidores."])
    instruction_sheet.append(["4_LINHAGEM_IMPORTACAO", "Declare relações explícitas source -> target quando desejar controle direto."])
    instruction_sheet.append(["5_LISTAS", "Referência rápida para tipos, camadas e relation types aceitos."])

    assets_sheet = workbook.create_sheet("2_ATIVOS")
    assets_sheet.append(LINEAGE_ASSET_HEADERS)
    assets = session.scalars(select(LineageAsset).order_by(LineageAsset.asset_name)).all()
    assets, assets_truncated = enforce_export_limit(assets, limit=limit)
    for asset in assets:
        assets_sheet.append(
            [
                redact_export_value(asset.asset_key, field_name="asset_key"),
                redact_export_value(asset.layer, field_name="layer"),
                redact_export_value(asset.asset_type, field_name="asset_type"),
                redact_export_value(asset.system_name, field_name="system_name"),
                redact_export_value(asset.schema_name, field_name="schema_name"),
                redact_export_value(asset.object_name, field_name="object_name"),
                redact_export_value(asset.asset_name, field_name="asset_name"),
                None,
                None,
                "Sim" if asset.is_active else "Não",
                redact_export_value(asset.description, field_name="description"),
                None,
            ]
        )

    mapping_sheet = workbook.create_sheet("3_MAPEAMENTO_NEGOCIO")
    mapping_sheet.append(LINEAGE_MAPPING_HEADERS)
    outgoing_by_source: dict[int, list[LineageRelation]] = defaultdict(list)
    incoming_by_target: dict[int, list[LineageRelation]] = defaultdict(list)
    active_relations = session.scalars(select(LineageRelation).where(LineageRelation.is_active.is_(True))).all()
    active_relations, relations_truncated = enforce_export_limit(active_relations, limit=limit)
    for relation in active_relations:
        outgoing_by_source[relation.source_asset_id].append(relation)
        incoming_by_target[relation.target_asset_id].append(relation)

    for asset in assets:
        if asset.asset_type == "source":
            continue
        incoming = incoming_by_target.get(asset.id, [])
        outgoing = outgoing_by_source.get(asset.id, [])
        upstream_keys = [relation.source_asset.asset_key for relation in incoming if relation.source_asset.asset_type != "source"]
        external_sources = [relation.source_asset for relation in incoming if relation.source_asset.asset_type == "source"]
        dashboards = [relation.target_asset.asset_name for relation in outgoing if relation.target_asset.asset_type in {"dashboard", "question"}]
        relation_with_process = next((relation for relation in [*incoming, *outgoing] if relation.process_name), None)
        safe_sheet_append(
            mapping_sheet,
            [
                asset.layer,
                asset.asset_type,
                asset.schema_name,
                asset.object_name,
                asset.asset_key,
                external_sources[0].system_name if external_sources else None,
                external_sources[0].asset_name if external_sources else None,
                "; ".join(upstream_keys) or None,
                redact_export_value(relation_with_process.process_name, field_name="process_name") if relation_with_process else None,
                redact_export_value(relation_with_process.process_type, field_name="process_type") if relation_with_process else None,
                "; ".join(dashboards) or None,
                redact_export_value(relation_with_process.notes, field_name="notes") if relation_with_process else None,
            ]
        )

    relations_sheet = workbook.create_sheet("4_LINHAGEM_IMPORTACAO")
    relations_sheet.append(LINEAGE_IMPORT_HEADERS)
    for relation in active_relations:
        relations_sheet.append(
            [
                redact_export_value(relation.source_asset.asset_key, field_name="source_asset_key"),
                redact_export_value(relation.target_asset.asset_key, field_name="target_asset_key"),
                redact_export_value(relation.relation_type, field_name="relation_type"),
                redact_export_value(relation.process_name, field_name="process_name"),
                redact_export_value(relation.process_type, field_name="process_type"),
                redact_export_value(relation.notes, field_name="notes"),
                "Sim" if relation.is_active else "Não",
            ]
        )

    lists_sheet = workbook.create_sheet("5_LISTAS")
    lists_sheet.append(["Lista", "Valores"])
    lists_sheet.append(["camada", "bronze; silver; gold; mart; dashboard; source; definir"])
    lists_sheet.append(["tipo_ativo", "table; view; dashboard; source"])
    lists_sheet.append(["relation_type", "ingestion; transformation; load; consumption"])

    header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    header_font = Font(bold=True, color="0F172A")
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        if sheet.max_row > 1 and sheet.max_column > 1 and sheet.title != "1_INSTRUCOES":
            last_column_letter = sheet.cell(row=1, column=sheet.max_column).column_letter
            table = Table(displayName=excel_table_name(sheet.title), ref=f"A1:{last_column_letter}{sheet.max_row}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            sheet.add_table(table)
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 36)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), {
        "assets_exported": len(assets),
        "assets_truncated": assets_truncated,
        "relations_exported": len(active_relations),
        "relations_truncated": relations_truncated,
    }


__all__ = ["build_lineage_workbook"]
