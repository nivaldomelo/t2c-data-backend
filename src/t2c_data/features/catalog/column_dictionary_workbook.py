from __future__ import annotations

from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass
from io import BytesIO
import re
import unicodedata

from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import Session

from t2c_data.features.export_security import redact_export_value
from t2c_data.features.tags.spreadsheet import TagSpreadsheetError, slugify_tag
from t2c_data.models.catalog import ColumnEntity, Schema, TableEntity
from t2c_data.features.audit import AuditFieldChange
from t2c_data.features.tags.intelligence import reprocess_table_tag_intelligence
from t2c_data.services.audit import log_field_changes
from t2c_data.schemas.catalog import ColumnDictionaryImportError, ColumnDictionaryImportResult

COLUMN_DICTIONARY_HEADERS = [
    "ID",
    "Slug",
    "Schema",
    "Tabela",
    "Posicao_Coluna",
    "Nome_Coluna",
    "Tipo_de_Dado",
    "UDT_Name",
    "Tamanho_Maximo",
    "Precisao_Numerica",
    "Escala_Numerica",
    "Aceita_Nulo",
    "Valor_Padrao",
    "Comentario_Existente",
    "Chave_Primaria",
    "Descricao",
    "Comentario",
]

COLUMN_DICTIONARY_HEADER_KEYS = []


def _normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_lookup_text(value: object) -> str | None:
    text = _normalize_text(value)
    if text is None:
        return None
    normalized = unicodedata.normalize("NFKD", text)
    stripped = "".join(char for char in normalized if not unicodedata.combining(char))
    collapsed = " ".join(stripped.split())
    return collapsed.casefold() or None


def _normalize_header_text(value: object) -> str:
    text = _normalize_lookup_text(value)
    if not text:
        return ""
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


COLUMN_DICTIONARY_HEADER_KEYS = [_normalize_header_text(header) for header in COLUMN_DICTIONARY_HEADERS]

@dataclass
class ParsedColumnDictionaryRow:
    row_number: int
    external_id: str | None
    slug: str
    schema_name: str
    table_name: str
    ordinal_position: int | None
    column_name: str
    data_type: str | None
    udt_name: str | None
    character_maximum_length: int | None
    numeric_precision: int | None
    numeric_scale: int | None
    is_nullable: bool | None
    column_default: str | None
    existing_comment: str | None
    is_primary_key: bool | None
    dictionary_description: str | None
    dictionary_comment: str | None


def _normalize_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text.replace(",", ".")))
    except ValueError as exc:  # pragma: no cover - validated for user payloads
        raise TagSpreadsheetError(f"Valor numérico inválido: {value}.") from exc


def _normalize_bool(value: object, field: str) -> bool | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return None
    truthy = {"sim", "s", "yes", "y", "true", "1", "x"}
    falsy = {"nao", "não", "n", "no", "false", "0"}
    if text in truthy:
        return True
    if text in falsy:
        return False
    raise TagSpreadsheetError(f"Valor inválido para {field}: {value}. Use Sim/Não.")


def _select_workbook_sheet(workbook):
    preferred_name = _normalize_header_text("Colunas_Importacao")
    for sheet in workbook.worksheets:
        if _normalize_header_text(sheet.title) == preferred_name:
            return sheet
    for sheet in workbook.worksheets:
        try:
            header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            continue
        header = [_normalize_header_text(value) for value in header_values]
        if header[: len(COLUMN_DICTIONARY_HEADER_KEYS)] == COLUMN_DICTIONARY_HEADER_KEYS:
            return sheet
    return workbook.active


def _iter_rows_from_workbook(content: bytes) -> Iterable[tuple[int, list[object | None]]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise TagSpreadsheetError("Dependência openpyxl não instalada no backend.") from exc

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise TagSpreadsheetError("Arquivo inválido. Envie uma planilha .xlsx compatível com o modelo.") from exc

    sheet = _select_workbook_sheet(workbook)
    try:
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration as exc:
        raise TagSpreadsheetError("A planilha não contém cabeçalho na primeira linha.") from exc
    header = [_normalize_header_text(value) for value in header_values]
    if header[: len(COLUMN_DICTIONARY_HEADER_KEYS)] != COLUMN_DICTIONARY_HEADER_KEYS:
        raise TagSpreadsheetError(
            "Cabeçalho inválido. Utilize o modelo com as colunas: " + ", ".join(COLUMN_DICTIONARY_HEADERS)
        )

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[: len(COLUMN_DICTIONARY_HEADERS)])
        if not any(value not in (None, "") for value in values):
            continue
        yield index, values


def parse_column_dictionary_workbook(
    content: bytes,
) -> tuple[list[ParsedColumnDictionaryRow], list[ColumnDictionaryImportError]]:
    parsed: list[ParsedColumnDictionaryRow] = []
    errors: list[ColumnDictionaryImportError] = []

    for row_number, values in _iter_rows_from_workbook(content):
        external_id = _normalize_text(values[0])
        raw_slug = _normalize_text(values[1])
        schema_name = _normalize_text(values[2])
        table_name = _normalize_text(values[3])
        ordinal_position = None
        character_maximum_length = None
        numeric_precision = None
        numeric_scale = None
        is_nullable = None
        is_primary_key = None
        column_name = _normalize_text(values[5])
        data_type = _normalize_text(values[6])
        udt_name = _normalize_text(values[7])
        column_default = _normalize_text(values[12])
        existing_comment = _normalize_text(values[13])
        dictionary_description = _normalize_text(values[15])
        dictionary_comment = _normalize_text(values[16])

        if not schema_name or not table_name or not column_name:
            errors.append(
                ColumnDictionaryImportError(
                    row_number=row_number,
                    slug=raw_slug,
                    message="Schema, Tabela e Nome_Coluna são obrigatórios.",
                )
            )
            continue

        try:
            ordinal_position = _normalize_int(values[4])
            character_maximum_length = _normalize_int(values[8])
            numeric_precision = _normalize_int(values[9])
            numeric_scale = _normalize_int(values[10])
            is_nullable = _normalize_bool(values[11], "Aceita_Nulo")
            is_primary_key = _normalize_bool(values[14], "Chave_Primaria")
        except TagSpreadsheetError as exc:
            errors.append(
                ColumnDictionaryImportError(
                    row_number=row_number,
                    slug=raw_slug,
                    message=str(exc),
                )
            )
            continue

        final_slug = slugify_tag(raw_slug or f"{schema_name}-{table_name}-{column_name}")
        if not final_slug:
            errors.append(
                ColumnDictionaryImportError(
                    row_number=row_number,
                    slug=raw_slug,
                    message="Slug inválido.",
                )
            )
            continue

        parsed.append(
            ParsedColumnDictionaryRow(
                row_number=row_number,
                external_id=external_id,
                slug=final_slug,
                schema_name=schema_name,
                table_name=table_name,
                ordinal_position=ordinal_position,
                column_name=column_name,
                data_type=data_type,
                udt_name=udt_name,
                character_maximum_length=character_maximum_length,
                numeric_precision=numeric_precision,
                numeric_scale=numeric_scale,
                is_nullable=is_nullable,
                column_default=column_default,
                existing_comment=existing_comment,
                is_primary_key=is_primary_key,
                dictionary_description=dictionary_description,
                dictionary_comment=dictionary_comment,
            )
        )

    return parsed, errors


@dataclass
class ExistingColumnLookup:
    by_external_id: dict[str, ColumnEntity | None]
    by_slug: dict[str, ColumnEntity | None]
    by_key: dict[tuple[str, str, str], ColumnEntity | None]


@dataclass
class ExistingTableLookup:
    by_key: dict[tuple[str, str], TableEntity | None]


@dataclass
class ColumnDictionaryCatalogGapSummary:
    duplicate_rows: int
    missing_catalog_rows: int
    missing_catalog_schemas: list[str]
    missing_catalog_tables: list[tuple[str, str, int]]


def _register_unique(mapping: dict, key, column: ColumnEntity) -> None:
    if key in mapping:
        if mapping[key] is not column:
            mapping[key] = None
        return
    mapping[key] = column


def _build_existing_lookup(session: Session, parsed_rows: list[ParsedColumnDictionaryRow]) -> ExistingColumnLookup:
    lookup_ids = sorted({normalized for row in parsed_rows if (normalized := _normalize_lookup_text(row.external_id))})
    lookup_slugs = sorted({normalized for row in parsed_rows if (normalized := _normalize_lookup_text(row.slug))})
    lookup_schemas = sorted({normalized for row in parsed_rows if (normalized := _normalize_lookup_text(row.schema_name))})
    lookup_tables = sorted({normalized for row in parsed_rows if (normalized := _normalize_lookup_text(row.table_name))})
    lookup_columns = sorted({normalized for row in parsed_rows if (normalized := _normalize_lookup_text(row.column_name))})

    conditions = []
    if lookup_ids:
        conditions.append(func.lower(func.trim(ColumnEntity.external_id)).in_(lookup_ids))
    if lookup_slugs:
        conditions.append(func.lower(func.trim(ColumnEntity.slug)).in_(lookup_slugs))
    if lookup_schemas and lookup_tables and lookup_columns:
        conditions.append(
            and_(
                func.lower(func.trim(Schema.name)).in_(lookup_schemas),
                func.lower(func.trim(TableEntity.name)).in_(lookup_tables),
                func.lower(func.trim(ColumnEntity.name)).in_(lookup_columns),
            )
        )

    if not conditions:
        return ExistingColumnLookup(by_external_id={}, by_slug={}, by_key={})

    existing_rows = session.execute(
        select(
            ColumnEntity,
            func.lower(func.trim(ColumnEntity.external_id)).label("external_id"),
            func.lower(func.trim(ColumnEntity.slug)).label("slug"),
            func.lower(func.trim(Schema.name)).label("schema_name"),
            func.lower(func.trim(TableEntity.name)).label("table_name"),
            func.lower(func.trim(ColumnEntity.name)).label("column_name"),
        )
        .join(TableEntity, ColumnEntity.table_id == TableEntity.id)
        .join(Schema, TableEntity.schema_id == Schema.id)
        .where(or_(*conditions))
    ).all()

    lookup = ExistingColumnLookup(by_external_id={}, by_slug={}, by_key={})
    for column, external_id, slug, schema_name, table_name, column_name in existing_rows:
        if external_id:
            _register_unique(lookup.by_external_id, external_id, column)
        if slug:
            _register_unique(lookup.by_slug, slug, column)
        key = (schema_name or "", table_name or "", column_name or "")
        _register_unique(lookup.by_key, key, column)
    return lookup


def _build_existing_table_lookup(session: Session, parsed_rows: list[ParsedColumnDictionaryRow]) -> ExistingTableLookup:
    lookup_keys = {
        (
            _normalize_lookup_text(row.schema_name) or "",
            _normalize_lookup_text(row.table_name) or "",
        )
        for row in parsed_rows
    }
    lookup_keys.discard(("", ""))
    if not lookup_keys:
        return ExistingTableLookup(by_key={})

    schemas = sorted({schema_name for schema_name, _ in lookup_keys if schema_name})
    tables = sorted({table_name for _, table_name in lookup_keys if table_name})
    existing_rows = session.execute(
        select(
            TableEntity,
            func.lower(func.trim(Schema.name)).label("schema_name"),
            func.lower(func.trim(TableEntity.name)).label("table_name"),
        )
        .join(Schema, TableEntity.schema_id == Schema.id)
        .where(
            func.lower(func.trim(Schema.name)).in_(schemas),
            func.lower(func.trim(TableEntity.name)).in_(tables),
        )
    ).all()

    lookup: ExistingTableLookup = ExistingTableLookup(by_key={})
    for table, schema_name, table_name in existing_rows:
        key = (schema_name or "", table_name or "")
        if key in lookup.by_key:
            if lookup.by_key[key] is not table:
                lookup.by_key[key] = None
            continue
        lookup.by_key[key] = table
    return lookup


def _resolve_existing_column(
    row: ParsedColumnDictionaryRow,
    lookup: ExistingColumnLookup,
) -> tuple[ColumnEntity | None, str | None, str | None]:
    ambiguity_reason: str | None = None

    external_id = _normalize_lookup_text(row.external_id)
    if external_id:
        if external_id in lookup.by_external_id:
            column = lookup.by_external_id[external_id]
            if column is None:
                ambiguity_reason = "ID ambiguo"
            else:
                return column, "ID", None

    slug = _normalize_lookup_text(row.slug)
    if slug:
        if slug in lookup.by_slug:
            column = lookup.by_slug[slug]
            if column is None:
                ambiguity_reason = ambiguity_reason or "Slug ambiguo"
            else:
                return column, "Slug", None

    key = (
        _normalize_lookup_text(row.schema_name) or "",
        _normalize_lookup_text(row.table_name) or "",
        _normalize_lookup_text(row.column_name) or "",
    )
    if key in lookup.by_key:
        column = lookup.by_key[key]
        if column is None:
            return None, None, ambiguity_reason or "Schema + Tabela + Nome_Coluna ambiguos"
        return column, "Schema + Tabela + Nome_Coluna", None

    return None, None, ambiguity_reason or "Coluna não encontrada no catálogo"


def _resolve_existing_table(
    row: ParsedColumnDictionaryRow,
    lookup: ExistingTableLookup,
) -> tuple[TableEntity | None, str | None]:
    key = (
        _normalize_lookup_text(row.schema_name) or "",
        _normalize_lookup_text(row.table_name) or "",
    )
    table = lookup.by_key.get(key)
    if table is None:
        if key in lookup.by_key:
            return None, "Tabela ambígua"
        return None, "Tabela não encontrada no catálogo"
    return table, "Schema + Tabela"


def _summarize_catalog_gaps(
    parsed_rows: list[ParsedColumnDictionaryRow],
    existing_lookup: ExistingColumnLookup,
    existing_table_lookup: ExistingTableLookup,
) -> ColumnDictionaryCatalogGapSummary:
    seen_keys: set[tuple[str, str, str]] = set()
    duplicate_rows = 0
    missing_catalog_rows = 0
    missing_schemas: set[str] = set()
    missing_tables: Counter[tuple[str, str]] = Counter()

    for row in parsed_rows:
        key = (row.schema_name.lower(), row.table_name.lower(), row.column_name.lower())
        if key in seen_keys:
            duplicate_rows += 1
            continue
        seen_keys.add(key)

        column, _, _ = _resolve_existing_column(row, existing_lookup)
        if column is not None:
            continue

        table, _ = _resolve_existing_table(row, existing_table_lookup)
        missing_catalog_rows += 1
        schema_key = row.schema_name.strip().lower()
        table_key = row.table_name.strip().lower()
        if schema_key:
            missing_schemas.add(schema_key)
        if table is None and schema_key and table_key:
            missing_tables[(schema_key, table_key)] += 1

    return ColumnDictionaryCatalogGapSummary(
        duplicate_rows=duplicate_rows,
        missing_catalog_rows=missing_catalog_rows,
        missing_catalog_schemas=sorted(missing_schemas),
        missing_catalog_tables=sorted(
            (
                schema_name,
                table_name,
                rows_count,
            )
            for (schema_name, table_name), rows_count in missing_tables.items()
        ),
    )


def _create_column_from_row(session: Session, row: ParsedColumnDictionaryRow, table: TableEntity) -> ColumnEntity:
    if row.ordinal_position is None:
        raise TagSpreadsheetError("Posicao_Coluna é obrigatória para criar uma nova coluna.")
    if not row.data_type:
        raise TagSpreadsheetError("Tipo_de_Dado é obrigatório para criar uma nova coluna.")
    column = ColumnEntity(
        table=table,
        name=row.column_name,
        data_type=row.data_type,
        is_primary_key=bool(row.is_primary_key),
        is_nullable=True if row.is_nullable is None else row.is_nullable,
        ordinal_position=row.ordinal_position,
        description_source=None,
        description_manual=None,
        external_id=row.external_id,
        slug=row.slug,
        udt_name=row.udt_name,
        character_maximum_length=row.character_maximum_length,
        numeric_precision=row.numeric_precision,
        numeric_scale=row.numeric_scale,
        column_default=row.column_default,
        existing_comment=row.existing_comment,
        dictionary_description=row.dictionary_description,
        dictionary_comment=row.dictionary_comment,
    )
    session.add(column)
    session.flush()
    return column


DICTIONARY_AUDIT_FIELDS = (
    "external_id",
    "slug",
    "udt_name",
    "character_maximum_length",
    "numeric_precision",
    "numeric_scale",
    "column_default",
    "existing_comment",
    "dictionary_description",
    "dictionary_comment",
)


def import_column_dictionary_from_workbook(
    session: Session,
    content: bytes,
    *,
    audit_kwargs: dict | None = None,
    actor_user_id: int | None = None,
    source_module: str = "catalog.column_dictionary",
    metadata: dict | None = None,
) -> ColumnDictionaryImportResult:
    parsed_rows, parsing_errors = parse_column_dictionary_workbook(content)
    errors = list(parsing_errors)
    imported = 0
    updated = 0
    matched = 0
    ignored = 0
    rejected = len(errors)
    touched_table_ids: set[int] = set()

    seen_keys: set[tuple[str, str, str]] = set()
    existing_lookup = _build_existing_lookup(session, parsed_rows)
    existing_table_lookup = _build_existing_table_lookup(session, parsed_rows)

    for row in parsed_rows:
        key = (row.schema_name.lower(), row.table_name.lower(), row.column_name.lower())
        if key in seen_keys:
            errors.append(
                ColumnDictionaryImportError(
                    row_number=row.row_number,
                    slug=row.slug,
                    message="Chave duplicada na planilha para Schema + Tabela + Nome_Coluna.",
                )
            )
            rejected += 1
            continue
        seen_keys.add(key)

        column, match_source, _ = _resolve_existing_column(row, existing_lookup)
        if column is None:
            table, table_match_source = _resolve_existing_table(row, existing_table_lookup)
            if table is None:
                errors.append(
                    ColumnDictionaryImportError(
                        row_number=row.row_number,
                        slug=row.slug,
                        message=(
                            "Tabela não encontrada no catálogo técnico atual. "
                            "Execute a sincronização do datasource antes de importar."
                        ),
                    )
                )
                rejected += 1
                continue
            match_source = table_match_source
            column = _create_column_from_row(session, row, table)
            before_snapshot = {field_name: None for field_name in DICTIONARY_AUDIT_FIELDS}
            after_snapshot = {field_name: getattr(column, field_name) for field_name in DICTIONARY_AUDIT_FIELDS}
            changes = [
                AuditFieldChange(field_name=field_name, before=before_snapshot[field_name], after=after_snapshot[field_name])
                for field_name in DICTIONARY_AUDIT_FIELDS
                if before_snapshot[field_name] != after_snapshot[field_name]
            ]
            matched += 1
            imported += 1
            touched_table_ids.add(int(table.id))
            created_column = True
        else:
            matched += 1
            before_snapshot = {field_name: getattr(column, field_name) for field_name in DICTIONARY_AUDIT_FIELDS}
            column.external_id = row.external_id
            column.slug = row.slug
            column.udt_name = row.udt_name
            column.character_maximum_length = row.character_maximum_length
            column.numeric_precision = row.numeric_precision
            column.numeric_scale = row.numeric_scale
            column.column_default = row.column_default
            column.existing_comment = row.existing_comment
            column.dictionary_description = row.dictionary_description
            column.dictionary_comment = row.dictionary_comment

            after_snapshot = {field_name: getattr(column, field_name) for field_name in DICTIONARY_AUDIT_FIELDS}
            changes = [
                AuditFieldChange(field_name=field_name, before=before_snapshot[field_name], after=after_snapshot[field_name])
                for field_name in DICTIONARY_AUDIT_FIELDS
                if before_snapshot[field_name] != after_snapshot[field_name]
            ]
            if not changes:
                ignored += 1
                continue
            updated += 1
            touched_table_ids.add(int(column.table_id))
            created_column = False

        if changes:
            log_field_changes(
                session,
                action="column_dictionary.import",
                entity_type="column",
                entity_id=column.id,
                parent_entity_type="table",
                parent_entity_id=column.table_id,
                changes=changes,
                source_module=source_module,
                metadata={
                    "row_number": row.row_number,
                    "schema_name": row.schema_name,
                    "table_name": row.table_name,
                    "column_name": row.column_name,
                    "match_source": match_source,
                    "created_column": created_column,
                    **(metadata or {}),
                },
                audit_kwargs=audit_kwargs,
                actor_user_id=actor_user_id,
            )

    for table_id in sorted(touched_table_ids):
        reprocess_table_tag_intelligence(
            session,
            table_id=table_id,
            actor_user_id=actor_user_id,
            audit_kwargs=audit_kwargs,
            source_module=source_module,
            metadata={"origin": "column_dictionary.import", **(metadata or {})},
        )
    session.commit()
    return ColumnDictionaryImportResult(
        processed=len(parsed_rows) + len(parsing_errors),
        matched=matched,
        imported=imported,
        updated=updated,
        ignored=ignored,
        rejected=rejected,
        errors=errors,
        touched_table_ids=sorted(touched_table_ids),
    )


def column_dictionary_export_rows(items: Iterable[tuple[ColumnEntity, str, str]]) -> list[list[str | int | None]]:
    rows: list[list[str | int | None]] = []
    for column, schema_name, table_name in items:
        rows.append(
            [
                redact_export_value(column.external_id, field_name="external_id"),
                redact_export_value(column.slug or slugify_tag(f"{schema_name}-{table_name}-{column.name}"), field_name="slug"),
                redact_export_value(schema_name, field_name="schema_name"),
                redact_export_value(table_name, field_name="table_name"),
                column.ordinal_position,
                redact_export_value(column.name, field_name="column_name"),
                redact_export_value(column.data_type, field_name="data_type"),
                redact_export_value(column.udt_name, field_name="udt_name"),
                column.character_maximum_length,
                column.numeric_precision,
                column.numeric_scale,
                "Sim" if column.is_nullable else "Não",
                redact_export_value(column.column_default, field_name="column_default"),
                redact_export_value(column.existing_comment, field_name="existing_comment"),
                "Sim" if column.is_primary_key else "Não",
                redact_export_value(
                    column.dictionary_description or column.description_manual or column.description_source,
                    field_name="dictionary_description",
                ),
                redact_export_value(column.dictionary_comment, field_name="dictionary_comment"),
            ]
        )
    return rows


def build_column_dictionary_workbook(
    items: Iterable[tuple[ColumnEntity, str, str]],
    *,
    include_readme: bool = True,
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise TagSpreadsheetError("Dependência openpyxl não instalada no backend.") from exc

    item_list = list(items)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Colunas_Importacao"
    sheet.append(COLUMN_DICTIONARY_HEADERS)
    header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    header_font = Font(bold=True, color="0F172A")

    for row in column_dictionary_export_rows(item_list):
        sheet.append(row)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    if sheet.max_row > 1:
        table_ref = f"A1:Q{sheet.max_row}"
        sheet_table = Table(displayName="ColumnDictionaryImport", ref=table_ref)
        sheet_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(sheet_table)

    widths = {
        "A": 14,
        "B": 26,
        "C": 20,
        "D": 24,
        "E": 16,
        "F": 26,
        "G": 20,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 16,
        "L": 14,
        "M": 20,
        "N": 28,
        "O": 16,
        "P": 36,
        "Q": 36,
    }
    for column_name, width in widths.items():
        sheet.column_dimensions[column_name].width = width

    if include_readme:
        readme = workbook.create_sheet("README")
        readme.append(["Campo", "Descrição"])
        for cell in readme[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in [
            ("ID", "Identificador externo prioritário para round-trip com a planilha."),
            ("Slug", "Chave estável secundária. Se vazia, é gerada a partir de Schema + Tabela + Nome_Coluna."),
            ("Schema/Tabela/Nome_Coluna", "Chave de fallback usada quando ID e Slug não estiverem disponíveis."),
            ("Descricao", "Descrição enriquecida exibida no Explorer."),
            ("Comentario", "Comentário adicional de negócio exibido no Explorer."),
            ("Comentario_Existente", "Comentário técnico já existente na base/origem."),
        ]:
            readme.append(row)
        readme.column_dimensions["A"].width = 28
        readme.column_dimensions["B"].width = 90

        summary = workbook.create_sheet("Resumo")
        summary.append(["Indicador", "Valor"])
        for cell in summary[1]:
            cell.fill = header_fill
            cell.font = header_font
        rows = list(column_dictionary_export_rows(item_list))
        summary.append(["Total de colunas exportadas", len(rows)])
        summary.append(["Com descrição enriquecida", sum(1 for row in rows if row[15])])
        summary.append(["Com comentário enriquecido", sum(1 for row in rows if row[16])])
        summary.column_dimensions["A"].width = 34
        summary.column_dimensions["B"].width = 18

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
