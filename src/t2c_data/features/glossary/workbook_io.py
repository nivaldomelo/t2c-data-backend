from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from io import BytesIO

from sqlalchemy import select
from sqlalchemy.orm import Session

from t2c_data.features.export_security import redact_export_value
from t2c_data.features.tags.spreadsheet import TagSpreadsheetError, slugify_tag
from t2c_data.models.glossary import GlossaryTerm
from t2c_data.models.tag import Tag
from t2c_data.schemas.glossary import (
    GlossarySpreadsheetImportError,
    GlossarySpreadsheetImportResult,
)

GLOSSARY_SPREADSHEET_HEADERS = [
    "ID",
    "Slug",
    "Termo",
    "Definicao",
    "Categoria",
    "Subcategoria",
    "Exemplo_de_Uso",
    "Sinonimos",
    "Prioridade_Sugerida",
    "Status",
    "Tags",
    "Observacoes",
]

ALLOWED_GLOSSARY_STATUS = {
    "ativo": "active",
    "active": "active",
    "inativo": "inactive",
    "inactive": "inactive",
    "rascunho": "draft",
    "draft": "draft",
    "deprecated": "deprecated",
    "descontinuado": "deprecated",
    "archived": "archived",
    "arquivado": "archived",
}

ALLOWED_PRIORITY = {
    "alta": "high",
    "high": "high",
    "media": "medium",
    "média": "medium",
    "medium": "medium",
    "baixa": "low",
    "low": "low",
}

STATUS_LABELS = {
    "active": "Ativo",
    "inactive": "Inativo",
    "draft": "Rascunho",
    "deprecated": "Descontinuado",
    "archived": "Arquivado",
}

PRIORITY_LABELS = {
    "high": "Alta",
    "medium": "Média",
    "low": "Baixa",
}


@dataclass
class ParsedGlossaryRow:
    row_number: int
    external_id: str | None
    slug: str
    name: str
    definition: str
    category: str | None
    subcategory: str | None
    example_of_use: str | None
    synonyms: str | None
    suggested_priority: str | None
    status: str
    tag_labels: str | None
    notes: str | None


def normalize_glossary_status(value: str | None) -> str:
    raw = (value or "active").strip().lower()
    if not raw:
        return "active"
    normalized = ALLOWED_GLOSSARY_STATUS.get(raw)
    if normalized is None:
        raise TagSpreadsheetError(
            f"Status inválido: {value}. Use um dos valores: {', '.join(sorted(STATUS_LABELS.values()))}."
        )
    return normalized


def normalize_priority(value: str | None) -> str | None:
    raw = (value or "").strip().lower()
    if not raw:
        return None
    normalized = ALLOWED_PRIORITY.get(raw)
    if normalized is None:
        raise TagSpreadsheetError(
            f"Prioridade inválida: {value}. Use um dos valores: {', '.join(sorted(PRIORITY_LABELS.values()))}."
        )
    return normalized


def glossary_status_label(value: str | None) -> str:
    return STATUS_LABELS.get(normalize_glossary_status(value), str(value or "Ativo"))


def glossary_priority_label(value: str | None) -> str:
    if not value:
        return ""
    normalized = normalize_priority(value)
    if not normalized:
        return ""
    return PRIORITY_LABELS.get(normalized, normalized.title())


def _normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _iter_rows_from_workbook(content: bytes):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise TagSpreadsheetError("Dependência openpyxl não instalada no backend.") from exc

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise TagSpreadsheetError("Arquivo inválido. Envie uma planilha .xlsx compatível com o modelo.") from exc

    sheet = workbook["Glossario_Importacao"] if "Glossario_Importacao" in workbook.sheetnames else workbook.active
    header = [
        str(value).strip() if value is not None else ""
        for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    if header[: len(GLOSSARY_SPREADSHEET_HEADERS)] != GLOSSARY_SPREADSHEET_HEADERS:
        raise TagSpreadsheetError(
            "Cabeçalho inválido. Utilize o modelo com as colunas: " + ", ".join(GLOSSARY_SPREADSHEET_HEADERS)
        )

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[: len(GLOSSARY_SPREADSHEET_HEADERS)])
        if not any(value not in (None, "") for value in values):
            continue
        yield index, values


def _normalize_tag_labels(session: Session, raw: str | None) -> str | None:
    if not raw:
        return None
    raw_items = [part.strip() for part in raw.replace(",", ";").split(";")]
    cleaned = [item for item in raw_items if item]
    if not cleaned:
        return None

    by_slug = {
        tag.slug.lower(): tag.slug for tag in session.scalars(select(Tag).where(Tag.slug.in_(cleaned))).all()
    }
    known_names = {
        tag.name.lower(): tag.slug
        for tag in session.scalars(select(Tag).where(Tag.name.in_(cleaned))).all()
    }

    normalized: list[str] = []
    for item in cleaned:
        key = item.lower()
        if key in by_slug:
            normalized.append(by_slug[key])
        elif key in known_names:
            normalized.append(known_names[key])
        else:
            normalized.append(slugify_tag(item))
    deduped: list[str] = []
    seen: set[str] = set()
    for item in normalized:
        if item in seen:
            continue
        seen.add(item)
        deduped.append(item)
    return ";".join(deduped)


def parse_glossary_workbook(
    session: Session, content: bytes
) -> tuple[list[ParsedGlossaryRow], list[GlossarySpreadsheetImportError]]:
    parsed: list[ParsedGlossaryRow] = []
    errors: list[GlossarySpreadsheetImportError] = []

    for row_number, values in _iter_rows_from_workbook(content):
        external_id = _normalize_text(values[0])
        slug = _normalize_text(values[1])
        name = _normalize_text(values[2])
        definition = _normalize_text(values[3])
        category = _normalize_text(values[4])
        subcategory = _normalize_text(values[5])
        example_of_use = _normalize_text(values[6])
        synonyms = _normalize_text(values[7])
        suggested_priority_value = _normalize_text(values[8])
        status_value = _normalize_text(values[9])
        tags_value = _normalize_text(values[10])
        notes = _normalize_text(values[11])

        if not name:
            errors.append(GlossarySpreadsheetImportError(row_number=row_number, slug=slug, message="Termo é obrigatório."))
            continue
        if not definition:
            errors.append(
                GlossarySpreadsheetImportError(row_number=row_number, slug=slug or name, message="Definição é obrigatória.")
            )
            continue

        final_slug = slugify_tag(slug or name)
        if not final_slug:
            errors.append(GlossarySpreadsheetImportError(row_number=row_number, slug=slug, message="Slug inválido."))
            continue

        try:
            normalized_status = normalize_glossary_status(status_value)
            normalized_priority = normalize_priority(suggested_priority_value)
            normalized_tags = _normalize_tag_labels(session, tags_value)
        except TagSpreadsheetError as exc:
            errors.append(GlossarySpreadsheetImportError(row_number=row_number, slug=final_slug, message=str(exc)))
            continue

        parsed.append(
            ParsedGlossaryRow(
                row_number=row_number,
                external_id=external_id,
                slug=final_slug,
                name=name,
                definition=definition,
                category=category,
                subcategory=subcategory,
                example_of_use=example_of_use,
                synonyms=synonyms,
                suggested_priority=normalized_priority,
                status=normalized_status,
                tag_labels=normalized_tags,
                notes=notes,
            )
        )

    return parsed, errors


def import_glossary_from_workbook(session: Session, content: bytes) -> GlossarySpreadsheetImportResult:
    parsed_rows, parsing_errors = parse_glossary_workbook(session, content)
    errors = list(parsing_errors)
    imported = 0
    updated = 0
    rejected = len(errors)

    seen_slugs: set[str] = set()
    seen_external_ids: set[str] = set()
    for row in parsed_rows:
        if row.slug in seen_slugs:
            errors.append(
                GlossarySpreadsheetImportError(row_number=row.row_number, slug=row.slug, message="Slug duplicado na planilha.")
            )
            rejected += 1
            continue
        seen_slugs.add(row.slug)

        if row.external_id:
            if row.external_id in seen_external_ids:
                errors.append(
                    GlossarySpreadsheetImportError(
                        row_number=row.row_number,
                        slug=row.slug,
                        message="ID duplicado na planilha.",
                    )
                )
                rejected += 1
                continue
            seen_external_ids.add(row.external_id)

        term = session.scalar(select(GlossaryTerm).where(GlossaryTerm.slug == row.slug))
        if term is None and row.external_id:
            term = session.scalar(select(GlossaryTerm).where(GlossaryTerm.external_id == row.external_id))

        if term is None:
            term = GlossaryTerm(
                external_id=row.external_id,
                slug=row.slug,
                name=row.name,
                definition=row.definition,
                description=row.definition,
                category=row.category,
                subcategory=row.subcategory,
                example_of_use=row.example_of_use,
                synonyms=row.synonyms,
                suggested_priority=row.suggested_priority,
                status=row.status,
                tag_labels=row.tag_labels,
                notes=row.notes,
            )
            session.add(term)
            imported += 1
            continue

        term.external_id = row.external_id or term.external_id
        term.slug = row.slug
        term.name = row.name
        term.definition = row.definition
        term.description = row.definition
        term.category = row.category
        term.subcategory = row.subcategory
        term.example_of_use = row.example_of_use
        term.synonyms = row.synonyms
        term.suggested_priority = row.suggested_priority
        term.status = row.status
        term.tag_labels = row.tag_labels
        term.notes = row.notes
        updated += 1

    session.commit()
    return GlossarySpreadsheetImportResult(
        processed=len(parsed_rows) + len(parsing_errors),
        imported=imported,
        updated=updated,
        rejected=rejected,
        errors=errors,
    )


def glossary_export_rows(items: Iterable[GlossaryTerm]) -> list[list[str]]:
    rows: list[list[str]] = []
    for term in items:
        rows.append(
            [
                redact_export_value(term.external_id, field_name="external_id"),
                redact_export_value(term.slug, field_name="slug"),
                redact_export_value(term.name, field_name="name"),
                redact_export_value(term.definition, field_name="definition"),
                redact_export_value(term.category, field_name="category"),
                redact_export_value(term.subcategory, field_name="subcategory"),
                redact_export_value(term.example_of_use, field_name="example_of_use"),
                redact_export_value(term.synonyms, field_name="synonyms"),
                redact_export_value(glossary_priority_label(term.suggested_priority), field_name="suggested_priority"),
                redact_export_value(glossary_status_label(term.status), field_name="status"),
                redact_export_value(term.tag_labels, field_name="tag_labels"),
                redact_export_value(term.notes, field_name="notes"),
            ]
        )
    return rows


def build_glossary_workbook(items: Iterable[GlossaryTerm], *, include_readme: bool = True) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise TagSpreadsheetError("Dependência openpyxl não instalada no backend.") from exc

    term_list = list(items)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Glossario_Importacao"
    sheet.append(GLOSSARY_SPREADSHEET_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="F97316")
    for row in glossary_export_rows(term_list):
        sheet.append(row)

    if sheet.max_row > 1:
        table = Table(displayName="tblGlossario", ref=f"A1:L{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)

    widths = [14, 24, 26, 56, 18, 18, 44, 24, 18, 14, 24, 24]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    if include_readme:
        readme = workbook.create_sheet("README")
        readme.append(["Campo", "Descrição"])
        for cell in readme[1]:
            cell.font = Font(bold=True)
        for row in [
            ("Slug", "Chave estável principal do glossário para importação e atualização."),
            ("Status", "Aceita Ativo, Inativo, Rascunho, Descontinuado e Arquivado."),
            ("Prioridade_Sugerida", "Aceita Alta, Média ou Baixa."),
            ("Tags", "Lista separada por ';'. Tags conhecidas são normalizadas por slug."),
        ]:
            readme.append(row)
        summary = workbook.create_sheet("Resumo")
        summary.append(["Indicador", "Valor"])
        summary.append(["Total de termos", len(term_list)])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
